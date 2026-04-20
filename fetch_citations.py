"""
Citation Count Fetcher
=======================
Reads an Excel file with a 'Publication_DOI' column and fetches the
citation count for each publication.

Strategy:
  1. OpenAlex (primary): query by DOI, extract cited_by_count.
     Batches up to 50 DOIs per request for efficiency.
  2. Crossref (fallback): for DOIs not found in OpenAlex, extract
     is-referenced-by-count from Crossref metadata.

Outputs the original Excel with added columns:
  - Citation_Count: total number of citations
  - Citation_Source: where the count came from (OpenAlex / Crossref)

Resilience: JSON cache, periodic saves, retry, graceful Ctrl+C.

Setup:
  Same config.ini as other scripts.

Usage:
  python fetch_citations.py <input.xlsx> [output.xlsx] [config.ini]
"""

import sys
import json
import time
import configparser
from pathlib import Path
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

OPENALEX_API = "https://api.openalex.org/works"
CROSSREF_API = "https://api.crossref.org/works/"
SCRIPT_DIR = Path(__file__).resolve().parent

BATCH_SIZE = 50  # OpenAlex supports up to 100 DOIs per filter query


# ── Config ──────────────────────────────────────────────────────────────

def load_config(config_path=None):
    if config_path is None:
        config_path = SCRIPT_DIR / "config.ini"
    else:
        config_path = Path(config_path).resolve()

    if not config_path.exists():
        print(f"ERROR: Config file not found at {config_path}")
        sys.exit(1)

    config = configparser.ConfigParser()
    config.read(config_path)
    email = config.get("crossref", "email", fallback=None)
    if not email or email.strip() == "your_email@example.com":
        print(f"ERROR: Please set your real email in {config_path}")
        sys.exit(1)

    return {
        "email": email.strip(),
        "delay": config.getfloat("crossref", "delay", fallback=1),
        "save_every": config.getint("crossref", "save_every", fallback=50),
        "max_retries": config.getint("crossref", "max_retries", fallback=3),
    }


# ── Cache ───────────────────────────────────────────────────────────────

def load_cache(cache_path):
    if cache_path.exists():
        with open(cache_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(cache, cache_path):
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


# ── OpenAlex batch fetch ───────────────────────────────────────────────

def fetch_openalex_batch(dois, session, email, max_retries=3):
    """
    Fetch citation counts for a batch of DOIs from OpenAlex.
    Uses the filter pipe syntax: doi:DOI1|DOI2|DOI3
    Returns dict of {doi: {cited_by_count, title, publication_year, source}}.
    """
    # Build filter: OpenAlex wants full DOI URLs with pipe separator
    doi_filter = "|".join(f"https://doi.org/{doi}" for doi in dois)

    params = {
        "filter": f"doi:{doi_filter}",
        "per_page": 100,
        "select": "doi,cited_by_count,title,publication_year,counts_by_year",
        "mailto": email,
    }

    for attempt in range(1, max_retries + 1):
        try:
            resp = session.get(OPENALEX_API, params=params, timeout=60)
            if resp.status_code == 429:
                wait = min(2 ** attempt * 5, 60)
                print(f"    OpenAlex rate limited. Waiting {wait}s...")
                time.sleep(wait)
                continue
            resp.raise_for_status()
            data = resp.json()

            results = {}
            for work in data.get("results", []):
                doi_url = work.get("doi", "")
                if doi_url:
                    # Normalise DOI: strip https://doi.org/ prefix
                    doi_clean = doi_url.replace("https://doi.org/", "").lower()
                    results[doi_clean] = {
                        "cited_by_count": work.get("cited_by_count", 0),
                        "title": work.get("title", ""),
                        "publication_year": work.get("publication_year"),
                        "counts_by_year": work.get("counts_by_year", []),
                        "source": "OpenAlex",
                    }

            return results

        except requests.exceptions.RequestException as e:
            if attempt < max_retries:
                time.sleep(2 ** attempt)
            else:
                print(f"    OpenAlex batch failed: {e}")
                return {}

    return {}


# ── Crossref single fetch ──────────────────────────────────────────────

def fetch_crossref_citation(doi, session, email, max_retries=3):
    """
    Fetch citation count for a single DOI from Crossref.
    Returns cited_by_count or None.
    """
    url = f"{CROSSREF_API}{doi}"
    params = {"mailto": email}

    for attempt in range(1, max_retries + 1):
        try:
            resp = session.get(url, params=params, timeout=30)
            if resp.status_code == 404:
                return None
            if resp.status_code == 429:
                time.sleep(min(2 ** attempt * 5, 60))
                continue
            resp.raise_for_status()
            data = resp.json()
            count = data.get("message", {}).get("is-referenced-by-count", 0)
            return count

        except requests.exceptions.RequestException as e:
            if attempt < max_retries:
                time.sleep(2 ** attempt)
            else:
                return None

    return None


# ── Main ────────────────────────────────────────────────────────────────

def main(input_file, output_file=None, config_path=None):
    cfg = load_config(config_path)
    print(f"Config: email={cfg['email']}, delay={cfg['delay']}s")

    input_path = Path(input_file)
    if not input_path.is_absolute():
        if not input_path.exists():
            fallback = SCRIPT_DIR / input_path
            if fallback.exists():
                input_path = fallback
    input_path = input_path.resolve()

    if not input_path.exists():
        print(f"ERROR: Input file not found: {input_path}")
        sys.exit(1)

    if output_file is None:
        output_path = input_path.parent / f"{input_path.stem}_with_citations.xlsx"
    else:
        output_path = Path(output_file).resolve()

    cache_path = output_path.parent / f"{input_path.stem}_citation_cache.json"

    print(f"Input:  {input_path}")
    print(f"Output: {output_path}")

    cache = load_cache(cache_path)
    if cache:
        print(f"  Loaded citation cache: {len(cache)} entries")

    # --- Read input ---
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    headers = {cell.value: cell.column for cell in ws[1]}
    doi_col = headers.get("Publication_DOI")

    if doi_col is None:
        print("ERROR: Could not find 'Publication_DOI' column.")
        sys.exit(1)

    # Collect unique DOIs
    doi_rows = {}
    for row in range(2, ws.max_row + 1):
        doi = ws.cell(row=row, column=doi_col).value
        if doi:
            doi = str(doi).strip()
            doi_rows.setdefault(doi, []).append(row)

    all_dois = list(doi_rows.keys())
    cached_dois = [d for d in all_dois if d.lower() in cache]
    uncached_dois = [d for d in all_dois if d.lower() not in cache]

    print(f"\nFound {len(all_dois)} unique DOIs across {sum(len(r) for r in doi_rows.values())} rows.")
    print(f"  Cached: {len(cached_dois)}, to fetch: {len(uncached_dois)}")

    # --- Phase 1: OpenAlex batch fetch ---
    session = requests.Session()
    session.headers.update({"User-Agent": f"CitationFetch/1.0 (mailto:{cfg['email']})"})

    if uncached_dois:
        print(f"\n--- Phase 1: OpenAlex (batches of {BATCH_SIZE}) ---")

        batches = [uncached_dois[i:i + BATCH_SIZE] for i in range(0, len(uncached_dois), BATCH_SIZE)]
        openalex_found = 0

        try:
            for batch_idx, batch in enumerate(batches):
                print(f"  Batch {batch_idx + 1}/{len(batches)} ({len(batch)} DOIs)")

                results = fetch_openalex_batch(batch, session, cfg["email"], cfg["max_retries"])

                for doi in batch:
                    doi_lower = doi.lower()
                    if doi_lower in results:
                        r = results[doi_lower]
                        cache[doi_lower] = {
                            "cited_by_count": r["cited_by_count"],
                            "source": "OpenAlex",
                        }
                        openalex_found += 1

                if (batch_idx + 1) % 5 == 0:
                    save_cache(cache, cache_path)
                    print(f"    >> Cache saved")

                time.sleep(cfg["delay"])

        except KeyboardInterrupt:
            print(f"\n>> Interrupted! Saving cache...")
            save_cache(cache, cache_path)
            print("Re-run to continue.")
            sys.exit(0)

        save_cache(cache, cache_path)
        print(f"  Found: {openalex_found}/{len(uncached_dois)}")

        # --- Phase 2: Crossref fallback for missing DOIs ---
        still_missing = [d for d in uncached_dois if d.lower() not in cache]

        if still_missing:
            print(f"\n--- Phase 2: Crossref fallback ({len(still_missing)} DOIs) ---")

            crossref_found = 0
            try:
                for i, doi in enumerate(still_missing):
                    display = doi[:60].encode("ascii", errors="replace").decode("ascii")
                    print(f"  [{i+1}/{len(still_missing)}] {display}")

                    count = fetch_crossref_citation(doi, session, cfg["email"], cfg["max_retries"])

                    if count is not None:
                        cache[doi.lower()] = {
                            "cited_by_count": count,
                            "source": "Crossref",
                        }
                        crossref_found += 1
                    else:
                        cache[doi.lower()] = {
                            "cited_by_count": None,
                            "source": "Not found",
                        }

                    if (i + 1) % cfg["save_every"] == 0:
                        save_cache(cache, cache_path)

                    time.sleep(cfg["delay"])

            except KeyboardInterrupt:
                print(f"\n>> Interrupted! Saving cache...")
                save_cache(cache, cache_path)
                print("Re-run to continue.")
                sys.exit(0)

            save_cache(cache, cache_path)
            print(f"  Found: {crossref_found}/{len(still_missing)}")

    # --- Write output ---
    # Add columns to the existing workbook
    # Check if Citation_Count column already exists
    if "Citation_Count" in headers:
        cite_col = headers["Citation_Count"]
        src_col = headers.get("Citation_Source", ws.max_column + 1)
    else:
        cite_col = ws.max_column + 1
        src_col = cite_col + 1
        ws.cell(row=1, column=cite_col, value="Citation_Count")
        ws.cell(row=1, column=src_col, value="Citation_Source")

    # Style the new headers
    header_font = Font(name="Arial", bold=True)
    ws.cell(row=1, column=cite_col).font = header_font
    ws.cell(row=1, column=src_col).font = header_font

    # Write citation data
    for doi, rows in doi_rows.items():
        doi_lower = doi.lower()
        cached_entry = cache.get(doi_lower, {})
        count = cached_entry.get("cited_by_count")
        source = cached_entry.get("source", "")

        for row in rows:
            ws.cell(row=row, column=cite_col, value=count if count is not None else "N/A")
            ws.cell(row=row, column=src_col, value=source)

    # Column widths
    ws.column_dimensions[openpyxl.utils.get_column_letter(cite_col)].width = 16
    ws.column_dimensions[openpyxl.utils.get_column_letter(src_col)].width = 16

    wb.save(output_path)

    # --- Stats ---
    total_with_citations = sum(1 for d in all_dois if cache.get(d.lower(), {}).get("cited_by_count") is not None)
    total_citations = sum(cache.get(d.lower(), {}).get("cited_by_count", 0) or 0 for d in all_dois)

    from collections import Counter
    source_counts = Counter(cache.get(d.lower(), {}).get("source", "Unknown") for d in all_dois)

    print(f"\n--- Summary ---")
    print(f"  DOIs with citation data: {total_with_citations}/{len(all_dois)}")
    print(f"  Total citations across all papers: {total_citations}")
    print(f"  By source:")
    for src in ["OpenAlex", "Crossref", "Not found", "Unknown"]:
        if source_counts.get(src, 0) > 0:
            print(f"    {src}: {source_counts[src]}")

    # Top cited papers
    cited_papers = []
    for doi in all_dois:
        entry = cache.get(doi.lower(), {})
        count = entry.get("cited_by_count")
        if count is not None and count > 0:
            cited_papers.append((doi, count))

    if cited_papers:
        cited_papers.sort(key=lambda x: x[1], reverse=True)
        print(f"\n  Top 10 most cited papers:")
        for doi, count in cited_papers[:10]:
            display = doi[:50].encode("ascii", errors="replace").decode("ascii")
            print(f"    {count:>6} citations | {display}")

    print(f"\nDone. Output saved to: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python fetch_citations.py <input.xlsx> [output.xlsx] [config.ini]")
        sys.exit(1)
    in_arg = sys.argv[1]
    out_arg = sys.argv[2] if len(sys.argv) > 2 else None
    cfg_arg = sys.argv[3] if len(sys.argv) > 3 else None
    main(in_arg, out_arg, cfg_arg)