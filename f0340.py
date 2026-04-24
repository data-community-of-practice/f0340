#!/usr/bin/env python3
"""
Connect Grants to Publications and Researchers
=================================================
Uses the publications Excel file as the bridge between grants,
publications (DOIs), and researchers (UUIDs).

Reads:
  - Publications_with_high_confidence.xlsx (Grant_ID <-> DOI mapping)
  - Researchers.json (researcher UUIDs + names)
  - Researcher_Publication.json (researcher UUID <-> DOI links)

Produces:
  - Grant_Publication.json — connects grant_id to publication DOI
  - Grant_Researcher.json — connects grant_id to researcher UUID

Usage:
  python f0340.py publications.xlsx Researchers.json Researcher_Publication.json
"""

import sys
import json
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).resolve().parent


def main():
    parser = argparse.ArgumentParser(
        description="Connect grants to publications and researchers"
    )
    parser.add_argument("publications_xlsx",
                        help="Publications Excel (with Grant_ID and Publication_DOI)")
    parser.add_argument("researchers_json",
                        help="Researchers.json (with id, full_name)")
    parser.add_argument("researcher_pub_json",
                        help="Researcher_Publication.json (researcher_id <-> publication_doi)")
    parser.add_argument("--output-dir", "-o", default=None,
                        help="Output directory (default: same as publications file)")
    args = parser.parse_args()

    pub_xlsx_path = Path(args.publications_xlsx).resolve()
    researchers_path = Path(args.researchers_json).resolve()
    res_pub_path = Path(args.researcher_pub_json).resolve()

    for p, label in [(pub_xlsx_path, "Publications xlsx"),
                     (researchers_path, "Researchers json"),
                     (res_pub_path, "Researcher_Publication json")]:
        if not p.exists():
            print(f"ERROR: {label} not found: {p}")
            sys.exit(1)

    output_dir = Path(args.output_dir).resolve() if args.output_dir else pub_xlsx_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)

    grant_pub_path = output_dir / "Grant_Publication.json"
    grant_res_path = output_dir / "Grant_Researcher.json"

    print(f"Publications xlsx:       {pub_xlsx_path}")
    print(f"Researchers json:        {researchers_path}")
    print(f"Researcher-Pub json:     {res_pub_path}")
    print(f"Output dir:              {output_dir}")
    print()

    # --- Read Grant_ID <-> DOI from Excel ---
    wb = openpyxl.load_workbook(pub_xlsx_path, read_only=True)
    ws = wb.active
    header = list(next(ws.iter_rows(max_row=1, values_only=True)))
    headers = {val: idx for idx, val in enumerate(header) if val is not None}

    grant_col = headers.get("Grant_ID")
    doi_col = headers.get("Publication_DOI")

    if grant_col is None or doi_col is None:
        print(f"ERROR: Need 'Grant_ID' and 'Publication_DOI' columns")
        print(f"  Found: {list(headers.keys())}")
        sys.exit(1)

    # Build grant <-> DOI mapping
    grant_pub_links = set()  # (grant_id, doi)
    for row in ws.iter_rows(min_row=2, values_only=True):
        grant_id = row[grant_col] if grant_col < len(row) else None
        doi = row[doi_col] if doi_col < len(row) else None

        if grant_id and doi:
            grant_pub_links.add((str(grant_id).strip(), str(doi).strip()))
    wb.close()

    print(f"Grant-Publication links:  {len(grant_pub_links)}")

    # --- Read Researcher_Publication relationships ---
    with open(res_pub_path, "r", encoding="utf-8") as f:
        res_pub_rels = json.load(f)

    # Build DOI -> set of researcher_ids
    doi_to_researchers = {}
    for rel in res_pub_rels:
        doi = rel["publication_doi"]
        rid = rel["researcher_id"]
        if doi not in doi_to_researchers:
            doi_to_researchers[doi] = set()
        doi_to_researchers[doi].add(rid)

    print(f"Researcher-Pub links:     {len(res_pub_rels)}")

    # --- Load researchers for name display ---
    with open(researchers_path, "r", encoding="utf-8") as f:
        researchers = json.load(f)
    researcher_names = {r["id"]: r.get("full_name", "") for r in researchers}

    print(f"Researchers:              {len(researchers)}")

    # --- Build Grant_Publication.json ---
    grant_pub_list = []
    seen_gp = set()
    for grant_id, doi in sorted(grant_pub_links):
        key = (grant_id, doi)
        if key not in seen_gp:
            seen_gp.add(key)
            grant_pub_list.append({
                "grant_id": grant_id,
                "publication_doi": doi,
            })

    # --- Build Grant_Researcher.json ---
    # A researcher is linked to a grant if they authored a publication
    # that is linked to that grant
    grant_res_links = set()  # (grant_id, researcher_id)
    for grant_id, doi in grant_pub_links:
        researcher_ids = doi_to_researchers.get(doi, set())
        for rid in researcher_ids:
            grant_res_links.add((grant_id, rid))

    grant_res_list = []
    for grant_id, rid in sorted(grant_res_links):
        grant_res_list.append({
            "grant_id": grant_id,
            "researcher_id": rid,
        })

    # --- Save ---
    with open(grant_pub_path, "w", encoding="utf-8") as f:
        json.dump(grant_pub_list, f, ensure_ascii=False, indent=2)

    with open(grant_res_path, "w", encoding="utf-8") as f:
        json.dump(grant_res_list, f, ensure_ascii=False, indent=2)

    # --- Summary ---
    unique_grants = set(g["grant_id"] for g in grant_pub_list)
    unique_dois = set(g["publication_doi"] for g in grant_pub_list)
    unique_grant_res = set(g["grant_id"] for g in grant_res_list)
    unique_res = set(g["researcher_id"] for g in grant_res_list)

    def safe(s):
        return s.encode("ascii", errors="replace").decode("ascii")

    print(f"\n{'='*55}")
    print(f"GRANT CONNECTION SUMMARY")
    print(f"{'='*55}")
    print(f"Grant_Publication.json:")
    print(f"  Links:              {len(grant_pub_list)}")
    print(f"  Unique grants:      {len(unique_grants)}")
    print(f"  Unique DOIs:        {len(unique_dois)}")
    print(f"Grant_Researcher.json:")
    print(f"  Links:              {len(grant_res_list)}")
    print(f"  Unique grants:      {len(unique_grant_res)}")
    print(f"  Unique researchers: {len(unique_res)}")
    print(f"{'='*55}")

    # Show a few examples
    print(f"\nSample grant-researcher links:")
    shown_grants = set()
    for rel in grant_res_list:
        gid = rel["grant_id"]
        if gid in shown_grants:
            continue
        shown_grants.add(gid)
        # Find all researchers for this grant
        grant_researchers = [r["researcher_id"] for r in grant_res_list if r["grant_id"] == gid]
        names = [safe(researcher_names.get(rid, "?")) for rid in grant_researchers[:5]]
        print(f"  Grant {gid}: {', '.join(names)}")
        if len(shown_grants) >= 5:
            break

    print(f"\nSaved:")
    print(f"  {grant_pub_path}")
    print(f"  {grant_res_path}")


if __name__ == "__main__":
    main()